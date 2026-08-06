#!/usr/bin/env python3
"""Point-in-time reconciliation of public sources against the Wind workbook."""

from __future__ import annotations

import argparse
import json
import math
from datetime import datetime, timedelta, timezone
from pathlib import Path

from openpyxl import load_workbook

import collect_macro_indicators as public


CN_TZ = timezone(timedelta(hours=8), name="Asia/Shanghai")


def month_key(value: str) -> str:
    return value[:7]


def lookup(dates: list[str], values: list[float], target: str, monthly: bool = False) -> tuple[str, float]:
    key = month_key(target) if monthly else target[:10]
    matches = [(date, value) for date, value in zip(dates, values) if (month_key(date) if monthly else date[:10]) == key]
    if not matches:
        raise KeyError(f"public series has no observation for {target}")
    return matches[-1]


def check(metric_id: str, definition: str, workbook_period: str, workbook_value: float,
          public_period: str, public_value: float, tolerance: float) -> dict:
    absolute = abs(float(public_value) - float(workbook_value))
    relative = absolute / max(abs(float(workbook_value)), 1e-12)
    return {
        "metric_id": metric_id,
        "definition": definition,
        "workbook_period": workbook_period,
        "public_period": public_period,
        "workbook_value": float(workbook_value),
        "public_value": float(public_value),
        "absolute_error": absolute,
        "relative_error_pct": 100 * relative,
        "tolerance_pct": 100 * tolerance,
        "status": "passed" if relative <= tolerance else "failed",
    }


def workbook_date(value) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    year = int(value)
    month = max(1, min(12, int(round((float(value) - year) * 100))))
    return f"{year:04d}-{month:02d}-01"


def validate(workbook: Path) -> dict:
    wb = load_workbook(workbook, data_only=True, read_only=True)
    now = datetime.now(CN_TZ)
    rows = []

    for metric_id, sheet_name, tickers in (
        ("macro_spy_djp", "SPYDJP", ("SPY", "DJP")),
        ("macro_discretionary_staples", "可选必选", ("^SP500-25", "^SP500-30")),
    ):
        ws = wb[sheet_name]
        target, reference = workbook_date(ws["A8"].value), float(ws["E8"].value)
        dates, values = public.aligned_ratio(public.yahoo_series(tickers[0]), public.yahoo_series(tickers[1]))
        date, value = lookup(dates, values, target)
        rows.append(check(metric_id, sheet_name, target, reference, date, value, 0.0001))

    ws = wb["DXY"]
    target, reference = workbook_date(ws["A8"].value), float(ws["B8"].value)
    date, value = lookup(*public.yahoo_series("DX-Y.NYB"), target)
    rows.append(check("macro_dxy", "美元指数收盘", target, reference, date, value, 0.0005))

    ws = wb["3个月商业票据利率"]
    target, reference = workbook_date(ws["A8"].value), float(ws["B8"].value)
    date, value = lookup(*public.parse_cp(), target)
    rows.append(check("macro_cp_rate", "AA非金融商业票据3个月", target, reference, date, value, 0.0001))

    ws = wb["LDR"]
    target, reference = workbook_date(ws["A8"].value), float(ws["M8"].value)
    dates, levels = public.parse_h8_ratio()
    transformed = public.weighted_roc_momentum(levels, (10, 13, 15, 20), (10, 13, 15, 20))
    valid = [(d, v) for d, v in zip(dates, transformed) if v is not None]
    date, value = lookup([x[0] for x in valid], [float(x[1]) for x in valid], target)
    rows.append(check("macro_ldr_kst", "对商业银行贷款/存款 KST", target, reference, date, value, 0.001))

    ws = wb["CPI"]
    target, reference = workbook_date(ws["A8"].value), float(ws["C8"].value)
    dates, levels = public.cpi_series(now)
    roc_dates, roc_values = dates[18:], [levels[i] / levels[i - 18] - 1 for i in range(18, len(levels))]
    date, value = lookup(roc_dates, roc_values, target, monthly=True)
    rows.append(check("macro_cpi_18roc", "季调CPI ROC18", target, reference, date, value, 0.0001))

    ws = wb["DIVIDEEND YIELD"]
    target, reference = workbook_date(ws["A8"].value), float(ws["L8"].value)
    dates, yields = public.multpl_series("s-p-500-dividend-yield")
    transformed = public.weighted_roc_momentum([1 / value for value in yields], (9, 12, 18, 24), (6, 6, 6, 9))
    valid = [(d, v) for d, v in zip(dates, transformed) if v is not None]
    date, value = lookup([x[0] for x in valid], [float(x[1]) for x in valid], target, monthly=True)
    rows.append(check("macro_dividend_kst", "1/股息率 KST", target, reference, date, value, 0.005))

    ws = wb["Margin Debt"]
    target, reference = workbook_date(ws["A8"].value), float(ws["K8"].value)
    dates, balances = public.finra_series()
    transformed = public.weighted_roc_momentum(balances, (9, 12, 18, 24), (6, 6, 6, 9))
    valid = [(d, v) for d, v in zip(dates, transformed) if v is not None]
    date, value = lookup([x[0] for x in valid], [float(x[1]) for x in valid], target, monthly=True)
    rows.append(check("macro_margin_kst", "FINRA保证金余额 KST", target, reference, date, value, 0.0001))

    ws = wb["ECY"]
    last_row = max(row for row in range(2, ws.max_row + 1) if isinstance(ws[f"B{row}"].value, (int, float)))
    target, reference = workbook_date(ws[f"A{last_row}"].value), float(ws[f"B{last_row}"].value)
    date, value = lookup(*public.shiller_ecy_series(), target, monthly=True)
    # Workbook ECY is stored to four decimal places; 0.5% relative tolerance
    # is tighter than the maximum rounding interval at the current level.
    rows.append(check("macro_excess_cape_yield", "Shiller Excess CAPE Yield", target, reference, date, value, 0.005))

    failed = [row["metric_id"] for row in rows if row["status"] != "passed"]
    return {
        "meta": {
            "validated_at": now.isoformat(timespec="seconds"),
            "workbook": workbook.name,
            "method": "same-period point reconciliation after workbook-defined transforms",
            "status": "passed" if not failed else "partial",
            "passed": len(rows) - len(failed),
            "total": len(rows),
            "failed_metrics": failed,
        },
        "checks": rows,
    }


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook")
    parser.add_argument("--project-root", default=".")
    parser.add_argument("--output", default="data/macro_source_validation.json")
    args = parser.parse_args()
    root = Path(args.project_root).resolve()
    workbook = Path(args.workbook)
    if not workbook.is_absolute():
        workbook = root / workbook
    result = validate(workbook)
    output = root / args.output
    output.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8", newline="\n")
    print(output)
    return 0 if result["meta"]["status"] == "passed" else 2


if __name__ == "__main__":
    raise SystemExit(main())
