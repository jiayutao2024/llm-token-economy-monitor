#!/usr/bin/env python3
"""Import an authorized Wind workbook as a non-redistributable CDF baseline.

The public repository stores only aggregated quantile knots and metadata, not
the proprietary dated observations from the workbook.
"""

from __future__ import annotations

import argparse
import json
import math
from datetime import datetime, timezone, timedelta
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string


CN_TZ = timezone(timedelta(hours=8), name="Asia/Shanghai")
SERIES = {
    "macro_spy_djp": ("SPYDJP", "E", "SPY.OF / DJP.OF", "PERCENTRANK.EXC"),
    "macro_discretionary_staples": ("可选必选", "E", "S5COND.SPI / S5CONS.SPI", "PERCENTRANK.EXC"),
    "macro_dxy": ("DXY", "B", "美元指数", "PERCENTRANK.EXC"),
    "macro_cp_rate": ("3个月商业票据利率", "B", "AA非金融商业票据3个月发行利率", "PERCENTRANK.EXC"),
    "macro_ldr_kst": ("LDR", "M", "对商业银行贷款/存款 KST", "ROC(10,13,15,20); SMA(10,13,15,20); W(1,2,3,4)"),
    "macro_cpi_18roc": ("CPI", "C", "CPI季调 ROC18", "CPI(t)/CPI(t-18)-1"),
    "macro_dividend_kst": ("DIVIDEEND YIELD", "L", "1/股息率 KST", "ROC(9,12,18,24); SMA(6,6,6,9); W(1,2,3,4)"),
    "macro_margin_kst": ("Margin Debt", "K", "保证金余额 KST", "ROC(9,12,18,24); SMA(6,6,6,9); W(1,2,3,4)"),
    "macro_excess_cape_yield": ("ECY", "B", "Shiller Excess CAPE Yield", "PERCENTRANK.EXC"),
}


def percentile_exc(sorted_values: list[float], probability: float) -> float:
    """Excel PERCENTILE.EXC-style interpolation with endpoint guards."""
    if probability <= 0:
        return sorted_values[0]
    if probability >= 1:
        return sorted_values[-1]
    rank = (len(sorted_values) + 1) * probability
    if rank <= 1:
        return sorted_values[0]
    if rank >= len(sorted_values):
        return sorted_values[-1]
    lower = math.floor(rank)
    fraction = rank - lower
    return sorted_values[lower - 1] + fraction * (sorted_values[lower] - sorted_values[lower - 1])


def quantile_knots(values: Iterable[float], steps: int = 100) -> list[list[float]]:
    clean = sorted(float(value) for value in values if isinstance(value, (int, float)) and math.isfinite(float(value)))
    if not clean:
        raise ValueError("baseline series has no numeric observations")
    return [[round(100 * i / steps, 3), percentile_exc(clean, i / steps)] for i in range(steps + 1)]


def extract(workbook: Path) -> dict:
    values = load_workbook(workbook, data_only=True, read_only=True)
    payload = {
        "meta": {
            "generated_at": datetime.now(CN_TZ).isoformat(timespec="seconds"),
            "workbook_name": workbook.name,
            "workbook_modified_at": datetime.fromtimestamp(workbook.stat().st_mtime, CN_TZ).isoformat(timespec="seconds"),
            "license_boundary": "Only aggregated 1-percentile CDF knots are retained; dated Wind observations are not published.",
            "percentile_method": "Excel PERCENTRANK.EXC-compatible interpolation against authorized historical baseline",
        },
        "series": {},
    }
    for metric_id, (sheet_name, column, definition, transform) in SERIES.items():
        sheet = values[sheet_name]
        observations = []
        periods = []
        column_index = column_index_from_string(column)
        for row in sheet.iter_rows(min_col=1, max_col=column_index, values_only=True):
            value = row[column_index - 1]
            if isinstance(value, (int, float)) and math.isfinite(float(value)):
                observations.append(float(value))
                raw_period = row[0]
                if isinstance(raw_period, datetime):
                    periods.append(raw_period.date().isoformat())
                elif sheet_name == "ECY" and isinstance(raw_period, (int, float)):
                    year = int(raw_period)
                    month = max(1, min(12, int(round((float(raw_period) - year) * 100))))
                    periods.append(f"{year:04d}-{month:02d}-01")
        if not observations:
            raise ValueError(f"{sheet_name}!{column} contains no numeric values")
        payload["series"][metric_id] = {
            "definition": definition,
            "transform": transform,
            "sample_count": len(observations),
            "min": min(observations),
            "max": max(observations),
            "quantile_knots": quantile_knots(observations),
            "source_sheet": sheet_name,
            "source_column": column,
            "latest_period": max(periods) if periods else None,
            "formula_anchor": transform,
        }
    return payload


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook")
    parser.add_argument("--project-root", default=".")
    parser.add_argument("--output", default="data/macro_baseline.json")
    args = parser.parse_args()
    root = Path(args.project_root).resolve()
    workbook = Path(args.workbook)
    if not workbook.is_absolute():
        workbook = root / workbook
    payload = extract(workbook)
    output = root / args.output
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8", newline="\n")
    print(output)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
